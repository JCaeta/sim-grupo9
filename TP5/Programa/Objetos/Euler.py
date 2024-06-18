import openpyxl

class Euler:
    def __init__(self, h):
        self.h = h

    def integracion_2000(self):
        integraciones = []

        cant_archivos = 2000
        a = cant_archivos
        t_siguiente = 0

        while a > 0:
            t = round(t_siguiente, 4)
            a = round(a, 4)

            da_dt = round(-68 - (a**2 / cant_archivos), 4)

            t_siguiente = round(t_siguiente + self.h, 4)
            a_euler = round((a + self.h * da_dt), 4)

            datos = {
                "t": t,
                "A": a,
                "dA_dt": da_dt,
                "t+1": t_siguiente,
                "A+1": a_euler
            }

            integraciones.append(datos)

            a = a_euler

        # llenar datos de tabla de excel

        fila_inicio = 16
        self.wb = openpyxl.load_workbook('../integraciones_euler.xlsx')
        self.ws = self.wb.active

        for idx, dato in enumerate(integraciones):
            self.ws[f"E{fila_inicio + idx}"] = dato["t"]
            self.ws[f"F{fila_inicio + idx}"] = dato["A"]
            self.ws[f"G{fila_inicio + idx}"] = dato["dA_dt"]
            self.ws[f"H{fila_inicio + idx}"] = dato["t+1"]
            self.ws[f"I{fila_inicio + idx}"] = dato["A+1"]

        self.wb.save('../integraciones_euler.xlsx')

    def integracion_1500(self):
        integraciones = []

        cant_archivos = 1500
        a = cant_archivos
        t_siguiente = 0

        while a > 0:
            t = round(t_siguiente, 4)
            a = round(a, 4)

            da_dt = round(-68 - (a**2 / cant_archivos), 4)

            t_siguiente = round(t_siguiente + self.h, 4)
            a_euler = round((a + self.h * da_dt), 4)

            datos = {
                "t": t,
                "A": a,
                "dA_dt": da_dt,
                "t+1": t_siguiente,
                "A+1": a_euler
            }

            integraciones.append(datos)

            a = a_euler

        # llenar datos de tabla de excel de ao = 1500

        fila_inicio = 16
        self.wb = openpyxl.load_workbook('../integraciones_euler.xlsx')
        self.ws = self.wb.active

        for idx, dato in enumerate(integraciones):
            self.ws[f"L{fila_inicio + idx}"] = dato["t"]
            self.ws[f"M{fila_inicio + idx}"] = dato["A"]
            self.ws[f"N{fila_inicio + idx}"] = dato["dA_dt"]
            self.ws[f"O{fila_inicio + idx}"] = dato["t+1"]
            self.ws[f"P{fila_inicio + idx}"] = dato["A+1"]

        self.wb.save('../integraciones_euler.xlsx')

    def integracion_1000(self):
        integraciones = []

        cant_archivos = 1000
        a = cant_archivos
        t_siguiente = 0

        while a > 0:
            t = round(t_siguiente, 4)
            a = round(a, 4)

            da_dt = round(-68 - (a**2 / cant_archivos), 4)

            t_siguiente = round(t_siguiente + self.h, 4)
            a_euler = round((a + self.h * da_dt), 4)

            datos = {
                "t": t,
                "A": a,
                "dA_dt": da_dt,
                "t+1": t_siguiente,
                "A+1": a_euler
            }

            integraciones.append(datos)

            a = a_euler

        # llenar datos de tabla de excel de ao = 1000

        fila_inicio = 16
        self.wb = openpyxl.load_workbook('../integraciones_euler.xlsx')
        self.ws = self.wb.active

        for idx, dato in enumerate(integraciones):
            self.ws[f"S{fila_inicio + idx}"] = dato["t"]
            self.ws[f"T{fila_inicio + idx}"] = dato["A"]
            self.ws[f"U{fila_inicio + idx}"] = dato["dA_dt"]
            self.ws[f"V{fila_inicio + idx}"] = dato["t+1"]
            self.ws[f"W{fila_inicio + idx}"] = dato["A+1"]

        self.wb.save('../integraciones_euler.xlsx')

    def calcular_tiempo(self, cant_archivos, numero_mantenimiento):
        self.wb = openpyxl.load_workbook('../integraciones_euler.xlsx')
        self.ws = self.wb.active
        fila = 16

        if cant_archivos == 2000:
            while self.ws[f"E{fila}"].value is not None:
                fila += 1

            ultimo_valor = round(self.ws[f"E{fila - 1}"].value, 4)
            nuevo_tiempo = round(ultimo_valor + self.h, 4)

            # Encontrar la primera fila vacía en la tabla de mantenimientos
            fila_mantenimiento = 16
            while self.ws[f"Y{fila_mantenimiento}"].value is not None:
                fila_mantenimiento += 1

            # Guardar los datos en la tabla de mantenimientos
            self.ws[f"Y{fila_mantenimiento}"] = numero_mantenimiento
            self.ws[f"Z{fila_mantenimiento}"] = cant_archivos
            self.ws[f"AA{fila_mantenimiento}"] = nuevo_tiempo
            self.wb.save('../integraciones_euler.xlsx')
            return nuevo_tiempo

        elif cant_archivos == 1500:
            while self.ws[f"L{fila}"].value is not None:
                fila += 1

            ultimo_valor = round(self.ws[f"L{fila - 1}"].value, 4)
            nuevo_tiempo = round(ultimo_valor + self.h, 4)

            # Encontrar la primera fila vacía en la tabla de mantenimientos
            fila_mantenimiento = 16
            while self.ws[f"Y{fila_mantenimiento}"].value is not None:
                fila_mantenimiento += 1

            # Guardar los datos en la tabla de mantenimientos
            self.ws[f"Y{fila_mantenimiento}"] = numero_mantenimiento
            self.ws[f"Z{fila_mantenimiento}"] = cant_archivos
            self.ws[f"AA{fila_mantenimiento}"] = nuevo_tiempo
            self.wb.save('../integraciones_euler.xlsx')
            return nuevo_tiempo

        elif cant_archivos == 1000:
            while self.ws[f"S{fila}"].value is not None:
                fila += 1

            ultimo_valor = round(self.ws[f"S{fila - 1}"].value, 4)
            nuevo_tiempo = round(ultimo_valor + self.h, 4)

            # Encontrar la primera fila vacía en la tabla de mantenimientos
            fila_mantenimiento = 16
            while self.ws[f"Y{fila_mantenimiento}"].value is not None:
                fila_mantenimiento += 1

            # Guardar los datos en la tabla de mantenimientos
            self.ws[f"Y{fila_mantenimiento}"] = numero_mantenimiento
            self.ws[f"Z{fila_mantenimiento}"] = cant_archivos
            self.ws[f"AA{fila_mantenimiento}"] = nuevo_tiempo

            self.wb.save('../integraciones_euler.xlsx')
            return nuevo_tiempo


