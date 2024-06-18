import tkinter as tk
from tkinter import messagebox
from logica import simulacion
from Objetos import Euler
import pandas as pd
from tksheet import Sheet
import openpyxl
from openpyxl.styles import PatternFill

class VentanaDatosFinales:
    def __init__(self, ac_tiempo_espera, ac_emp_pasaron, ac_emp_salen, reloj):
        self.ventana = tk.Tk()
        self.ventana.title("Tabla de Simulación")
        self.ventana.title("TP Nº5 - SIMULACIÓN - 4K3 - UTN FRC - GRUPO 9")
        self.ventana.geometry("700x100")
        self.ventana.configure(bg="#f3f3d1")

        self.porcentaje_empleados_que_se_van_temporalmente = str(round(float(ac_emp_salen * 100 / ac_emp_pasaron), 2)) + ' %'
        self.tiempo_promedio_espera = str(round(float(ac_tiempo_espera / reloj), 2)) + ' minutos'

        self.label_dato1 = tk.Label(self.ventana,
                                    text=f"Porcentaje de empleados que se van temporalmente para volver mas tarde: "
                                         f"{self.porcentaje_empleados_que_se_van_temporalmente}",
                                    font=("Helvetica", 12, "bold"),
                                    bg='#f3f3d1')

        self.label_dato1.place(x=30, y=30)

        self.label_dato2 = tk.Label(self.ventana,
                                    text=f"Tiempo promedio de espera: "
                                         f"{self.tiempo_promedio_espera}",
                                    font=("Helvetica", 12, "bold"),
                                    bg='#f3f3d1')
        self.label_dato2.place(x=30, y=60)


class TablaPandas:
    def __init__(self, simulaciones, ac_tiempo_espera, ac_emp_pasaron, ac_emp_salen, reloj):
        self.datos = [sim.to_dict() for sim in simulaciones]

        df = pd.DataFrame(self.datos)

        root = tk.Tk()
        root.title("Tabla de Simulación")

        sheet = Sheet(root,
                      data=df.values.tolist(),  # Convertir el DataFrame a una lista de listas
                      headers=list(df.columns),  # Usar los nombres de las columnas del DataFrame como encabezados
                      width=1000,
                      height=400,
                      column_width=200)

        sheet.enable_bindings("single_select",
                              "row_select",
                              "column_width_resize",
                              "arrowkeys",
                              "right_click_popup_menu",
                              "rc_select",
                              "copy",
                              "cut",
                              "paste",
                              "delete",
                              "undo")
        for _ in df.columns:
            sheet.align(align="center")

        # Empaquetar el widget Sheet
        sheet.pack(expand=True, fill='both')

        VentanaDatosFinales(ac_tiempo_espera, ac_emp_pasaron, ac_emp_salen, reloj)

class ExcelEuler:
    def __init__(self, h):
        self.h = h
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def set_estructura(self):
        # Definir el contenido y la estructura
        data = [
            ("G10", "Ao"), ("H10", 2000),

            ("G11", "h"), ("H11", self.h),

            ("N10", "Ao"), ("O10", 1500),

            ("N11", "h"), ("O11", self.h),

            ("U10", "Ao"), ("V10", 1000),

            ("U11", "h"), ("V11", self.h),

            # Encabezados
            ("E15", "t"), ("F15", "A"), ("G15", "dA/dt"), ("H15", "t+1"), ("I15", "A+1 e"),
            ("L15", "t"), ("M15", "A"), ("N15", "dA/dt"), ("O15", "t+1"), ("P15", "A+1 e"),
            ("S15", "t"), ("T15", "A"), ("U15", "dA/dt"), ("V15", "t+1"), ("W15", "A+1 e"),

            # Datos
            ("H16", 0), ("I16", 2000),
            ("O16", 0), ("P16", 1500),
            ("V16", 0), ("W16", 1000),

            # Probabilidades
            ("Y5", "Cant. Archivos"), ("Z5", "P()"), ("AA5", "P() AC"),
            ("Y6", 1000), ("Z6", "0,33"), ("AA6", "0.33"),
            ("Y7", 1500), ("Z7", "0,33"), ("AA7", "0.67"),
            ("Y8", 2000), ("Z8", "0,33"), ("AA8", "1"),

            # Fila 21 - Encabezados tabla final
            ("Y15", "Mantenimiento n°"),
            ("Z15", "Cant. Archivos"),
            ("AA15", "Tiempo (minutos)")
        ]

        # Aplicar el contenido y la estructura
        for cell, value in data:
            self.ws[cell] = value

        # Aplicar fondo rojo a las filas de encabezados
        fill = PatternFill(start_color="EAB676", end_color="EAB676", fill_type="solid")
        for cell in ["E15", "F15", "G15", "H15", "I15", "L15", "M15", "N15", "O15", "P15", "S15", "T15", "U15", "V15", "W15", "Y15", "Z15", "AA15"]:
            self.ws[cell].fill = fill

        self.ws.column_dimensions['Y'].width = 18
        self.ws.column_dimensions['Z'].width = 18
        self.ws.column_dimensions['AA'].width = 20

    def guardar(self, filename):
        self.wb.save(filename)


class App:
    def __init__(self):
        self.ventana = tk.Tk()
        self.ventana.title("TP Nº5 - SIMULACIÓN - 4K3 - UTN FRC - GRUPO 9")
        self.ventana.geometry("1000x320")
        self.ventana.configure(bg="#f3f3d1")

        icon = tk.PhotoImage(file='utn.png')
        self.ventana.iconphoto(False, icon)

        self.label_cantidad_tiempo = tk.Label(text="Cantidad de tiempo (X)", font=("Helvetica", 14, "bold"),
                                              bg='#f3f3d1')
        self.label_cantidad_tiempo.place(x=20, y=20)
        self.entry_cantidad_tiempo = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_cantidad_tiempo.place(x=65, y=50)

        self.label_cantidad_iteraciones_a_mostrar = tk.Label(text="Cant. Iteraciones a mostrar (i)",
                                                             font=("Helvetica", 14, "bold"), bg='#f3f3d1')
        self.label_cantidad_iteraciones_a_mostrar.place(x=10, y=80)
        self.entry_cantidad_iteraciones_a_mostrar = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_cantidad_iteraciones_a_mostrar.place(x=65, y=110)

        self.label_minuto_desde = tk.Label(text="Minuto desde (j)", font=("Helvetica", 14, "bold"), bg='#f3f3d1')
        self.label_minuto_desde.place(x=50, y=140)
        self.entry_minuto_desde = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_minuto_desde.place(x=65, y=170)

        self.label_h = tk.Label(text="Valor de h", font=("Helvetica", 14, "bold"), bg='#f3f3d1')
        self.label_h.place(x=65, y=200)
        self.entry_h = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_h.place(x=65, y=230)

        # MEDIO DE LA VENTANA #
        self.label_evento_llegada_empleado = tk.Label(text="Evento llegada_empleado", font=("Helvetica", 14, "bold"),
                                                      bg='#f3f3d1')
        self.label_evento_llegada_empleado.place(x=300, y=20)
        self.entry_media_evento_llegada_empleado = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_media_evento_llegada_empleado.place(x=350, y=50)
        self.label_media_llegada_empleado = tk.Label(text="Media", font=("Helvetica", 12), bg='#f3f3d1')
        self.label_media_llegada_empleado.place(x=460, y=50)

        self.label_evento_llegada_tecnico = tk.Label(text="Evento llegada_tecnico", font=("Helvetica", 14, "bold"),
                                                     bg='#f3f3d1')
        self.label_evento_llegada_tecnico.place(x=300, y=110)
        self.label_hora_A_llegada_tecnico = tk.Label(text="A (en horas)", font=("Helvetica", 12), bg='#f3f3d1')
        self.label_hora_A_llegada_tecnico.place(x=460, y=140)
        self.entry_hora_A_llegada_tecnico = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_hora_A_llegada_tecnico.place(x=350, y=140)
        self.label_minuto_B_llegada_tecnico = tk.Label(text="B (en minutos)", font=("Helvetica", 12), bg='#f3f3d1')
        self.label_minuto_B_llegada_tecnico.place(x=460, y=170)
        self.entry_minuto_B_llegada_tecnico = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_minuto_B_llegada_tecnico.place(x=350, y=170)

        # DERECHO DE LA VENTANA #
        self.label_evento_fin_registro_huella = tk.Label(text="Evento fin_registro_huella",
                                                         font=("Helvetica", 14, "bold"), bg='#f3f3d1')
        self.label_evento_fin_registro_huella.place(x=600, y=20)
        self.entry_minuto_A_fin_registro_huella = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_minuto_A_fin_registro_huella.place(x=650, y=50)
        self.label_minuto_A_fin_registro_huella = tk.Label(text="A (en minutos)", font=("Helvetica", 12), bg='#f3f3d1')
        self.label_minuto_A_fin_registro_huella.place(x=750, y=50)
        self.entry_minuto_B_fin_registro_huella = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_minuto_B_fin_registro_huella.place(x=650, y=80)
        self.label_minuto_B_fin_registro_huella = tk.Label(text="B (en minutos)", font=("Helvetica", 12), bg='#f3f3d1')
        self.label_minuto_B_fin_registro_huella.place(x=750, y=80)

        self.label_evento_fin_mantenimiento_terminal = tk.Label(text="Evento fin_mantenimiento_terminal",
                                                                font=("Helvetica", 14, "bold"), bg='#f3f3d1')
        self.label_evento_fin_mantenimiento_terminal.place(x=600, y=110)
        self.entry_minuto_A_fin_mantenimiento_terminal = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_minuto_A_fin_mantenimiento_terminal.place(x=650, y=140)
        self.label_minuto_A_fin_mantenimiento_terminal = tk.Label(text="A (en minutos)", font=("Helvetica", 12),
                                                                  bg='#f3f3d1')
        self.label_minuto_A_fin_mantenimiento_terminal.place(x=750, y=140)
        self.entry_minuto_B_fin_mantenimiento_terminal = tk.Entry(font=("Helvetica", 12), width=10)
        self.entry_minuto_B_fin_mantenimiento_terminal.place(x=650, y=170)
        self.label_minuto_B_fin_mantenimiento_terminal = tk.Label(text="B (en minutos)", font=("Helvetica", 12),
                                                                  bg='#f3f3d1')
        self.label_minuto_B_fin_mantenimiento_terminal.place(x=750, y=170)

        self.button = tk.Button(text="Generar simulación", font=("Consolas", 16, "bold"),
                                bg="lightblue", fg="black", command=self.validacion_datos)
        self.button.place(x=360, y=250, width=300, height=50)

    def validacion_datos(self):
        try:
            cantidad_tiempo = float(self.entry_cantidad_tiempo.get())
            cantidad_iteraciones_a_mostrar = int(self.entry_cantidad_iteraciones_a_mostrar.get())
            minuto_desde = int(self.entry_minuto_desde.get())
            media_llegada_empleado = float(self.entry_media_evento_llegada_empleado.get())
            hora_A_llegada_tecnico = float(self.entry_hora_A_llegada_tecnico.get())
            minuto_B_llegada_tecnico = float(self.entry_minuto_B_llegada_tecnico.get())
            minuto_A_fin_registro_huella = float(self.entry_minuto_A_fin_registro_huella.get())
            minuto_B_fin_registro_huella = float(self.entry_minuto_B_fin_registro_huella.get())
            minuto_A_fin_mantenimiento_terminal = float(self.entry_minuto_A_fin_mantenimiento_terminal.get())
            minuto_B_fin_mantenimiento_terminal = float(self.entry_minuto_B_fin_mantenimiento_terminal.get())
            valor_h = float(self.entry_h.get())

            minuto_B_es_correcto = True if (hora_A_llegada_tecnico * 60 - minuto_B_llegada_tecnico >= 0) else False

            if cantidad_tiempo > 0 and(cantidad_iteraciones_a_mostrar >= 0 or cantidad_iteraciones_a_mostrar > 100000) and (minuto_desde >= 0):
                if (media_llegada_empleado > 0) and valor_h > 0:
                    if (hora_A_llegada_tecnico >= 0 and minuto_B_es_correcto):
                        if (minuto_A_fin_registro_huella >= 0 and minuto_B_fin_registro_huella >= 0 and minuto_B_fin_registro_huella > minuto_A_fin_registro_huella):
                            if (minuto_A_fin_mantenimiento_terminal >= 0 and minuto_B_fin_mantenimiento_terminal >= 0 and minuto_B_fin_mantenimiento_terminal > minuto_A_fin_mantenimiento_terminal):

                                excel = ExcelEuler(h=valor_h)
                                excel.set_estructura()
                                excel.guardar('../integraciones_euler.xlsx')

                                euler = Euler.Euler(h=valor_h)
                                euler.integracion_2000()
                                euler.integracion_1500()
                                euler.integracion_1000()

                                datos, ac_tiempo_espera, ac_emp_pasaron, ac_emp_salen, reloj = simulacion(
                                    minutoARegistroHuella=minuto_A_fin_registro_huella,
                                    minutoBRegistroHuella=minuto_B_fin_registro_huella,
                                    mediaLlegadaEmpleado=media_llegada_empleado,
                                    horaALlegadaTecnico=hora_A_llegada_tecnico,
                                    minutoBLlegadaTecnico=minuto_B_llegada_tecnico,
                                    minutoAMantenimientoTerminal=minuto_A_fin_mantenimiento_terminal,
                                    minutoBMantenimientoTerminal=minuto_B_fin_mantenimiento_terminal,
                                    cantidad_tiempo=cantidad_tiempo,
                                    cantidad_datos_a_mostrar=cantidad_iteraciones_a_mostrar,
                                    minuto_desde=minuto_desde, h=valor_h)
                                TablaPandas(datos, ac_tiempo_espera, ac_emp_pasaron, ac_emp_salen, reloj)


                            else:
                                messagebox.showerror("Error", "Ha ingresado algún dato erróneo. Revise de vuelta")
                                return
                        else:
                            messagebox.showerror("Error", "Ha ingresado algún dato erróneo. Revise de vuelta")
                            return
                    else:
                        messagebox.showerror("Error", "Ha ingresado algún dato erróneo. Revise de vuelta")
                        return
                else:
                    messagebox.showerror("Error", "Ha ingresado algún dato erróneo. Revise de vuelta")
                    return
            else:
                messagebox.showerror("Error", "Ha ingresado algún dato erróneo. Revise de vuelta")
                return

        except ValueError:
            messagebox.showerror("Error", "Se detectó un ingreso de caracter inválido.")
